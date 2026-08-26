import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, date, timedelta

import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from flights import search_prices_by_airline
from discounts import get_discounts_for_airline

# notifications.py
#
# This is the "manual tracker" feature: instead of just searching once,
# a student can ask us to watch a route on a few specific airlines for a
# handful of days, and we'll email them a proper comparison once that
# window's up -- basically doing the tedious job of checking prices every
# day yourself, which is what a lot of people already do by hand in a
# spreadsheet.
#
# Two tables live in the same Postgres database as price_history:
#   price_trackers        -- one row per subscription someone set up
#   tracker_daily_prices   -- one row per (tracker, day, airline) price we recorded
#
# The actual "check today's prices and see if any tracker is due" logic
# lives in poller.py, since that's the process that's already running
# continuously in the background -- this file just provides the building
# blocks it calls.

_DATABASE_URL = os.getenv("DATABASE_URL")
_GMAIL_ADDRESS = os.getenv("GMAIL_ADDRESS")
_GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD")

# Colours pulled straight from the frontend's palette, so a tracker report
# and the website itself feel like the same product rather than a plain
# spreadsheet (or plain email) bolted on the side.
NAVY = "14213D"
GOLD = "C98F1F"
CREAM = "FDF8F0"
CREAM_DIM = "F3ECDC"
SAGE = "D9F0E6"
WHITE = "FFFFFF"


def _get_connection():
    return psycopg2.connect(_DATABASE_URL)


def init_notifications_db():
    """
    Creates both tracker tables if they don't already exist. Safe to call
    every time the app starts -- CREATE TABLE IF NOT EXISTS just does
    nothing on the second and every later run.
    """
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS price_trackers (
                    id SERIAL PRIMARY KEY,
                    email TEXT NOT NULL,
                    origin TEXT NOT NULL,
                    destination TEXT NOT NULL,
                    seat_class TEXT NOT NULL DEFAULT 'economy',
                    airlines TEXT NOT NULL,
                    start_date DATE NOT NULL,
                    duration_days INTEGER NOT NULL,
                    active BOOLEAN NOT NULL DEFAULT TRUE,
                    created_at TIMESTAMPTZ NOT NULL
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS tracker_daily_prices (
                    id SERIAL PRIMARY KEY,
                    tracker_id INTEGER NOT NULL REFERENCES price_trackers(id),
                    checked_date DATE NOT NULL,
                    airline TEXT NOT NULL,
                    price INTEGER,
                    UNIQUE (tracker_id, checked_date, airline)
                )
            """)
        conn.commit()
    finally:
        conn.close()


def create_tracker(email: str, origin: str, destination: str, airlines: list[str], duration_days: int, seat_class: str = "economy") -> int:
    """Registers a new tracker and returns its id."""
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO price_trackers (email, origin, destination, seat_class, airlines, start_date, duration_days, active, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, TRUE, %s)
                RETURNING id
                """,
                (email, origin, destination, seat_class, ",".join(airlines), date.today(), duration_days, datetime.now()),
            )
            new_id = cur.fetchone()[0]
        conn.commit()
        return new_id
    finally:
        conn.close()


def get_due_trackers() -> list[dict]:
    """Trackers that are still active and inside their tracking window today."""
    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, email, origin, destination, seat_class, airlines, start_date, duration_days
                FROM price_trackers
                WHERE active = TRUE
                  AND CURRENT_DATE >= start_date
                  AND CURRENT_DATE < start_date + duration_days
                """
            )
            return cur.fetchall()
    finally:
        conn.close()


def get_trackers_ready_to_close() -> list[dict]:
    """Trackers whose window has fully passed -- time to send the report."""
    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, email, origin, destination, seat_class, airlines, start_date, duration_days
                FROM price_trackers
                WHERE active = TRUE
                  AND CURRENT_DATE >= start_date + duration_days
                """
            )
            return cur.fetchall()
    finally:
        conn.close()


def has_checked_today(tracker_id: int) -> bool:
    """Stops the poller (which runs every 6 hours) from writing 4 rows for
    the same day instead of 1."""
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT COUNT(*) FROM tracker_daily_prices WHERE tracker_id = %s AND checked_date = CURRENT_DATE",
                (tracker_id,),
            )
            return cur.fetchone()[0] > 0
    finally:
        conn.close()


def save_tracker_day(tracker_id: int, prices_by_airline: dict):
    """Saves one day's worth of prices for a tracker, one row per airline."""
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            for airline, price in prices_by_airline.items():
                cur.execute(
                    """
                    INSERT INTO tracker_daily_prices (tracker_id, checked_date, airline, price)
                    VALUES (%s, CURRENT_DATE, %s, %s)
                    ON CONFLICT (tracker_id, checked_date, airline) DO NOTHING
                    """,
                    (tracker_id, airline, price),
                )
        conn.commit()
    finally:
        conn.close()


def get_tracker_history(tracker_id: int) -> list[dict]:
    """Every price point recorded for a tracker, oldest day first."""
    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                "SELECT checked_date, airline, price FROM tracker_daily_prices WHERE tracker_id = %s ORDER BY checked_date ASC",
                (tracker_id,),
            )
            return cur.fetchall()
    finally:
        conn.close()


def close_tracker(tracker_id: int):
    """Marks a tracker inactive once its report's been sent."""
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("UPDATE price_trackers SET active = FALSE WHERE id = %s", (tracker_id,))
        conn.commit()
    finally:
        conn.close()


def _build_report_excel(tracker: dict, history: list[dict], filepath: str):
    """Builds the .xlsx report -- Summary + Daily Prices sheets, styled to
    match Stufly's own colours."""
    airlines = tracker["airlines"].split(",")
    origin, destination = tracker["origin"], tracker["destination"]

    by_date = {}
    for row in history:
        d = row["checked_date"]
        by_date.setdefault(d, {})[row["airline"]] = row["price"]
    sorted_dates = sorted(by_date.keys())

    best_price, best_airline, best_date = None, None, None
    for d in sorted_dates:
        for airline, price in by_date[d].items():
            if price is not None and (best_price is None or price < best_price):
                best_price, best_airline, best_date = price, airline, d

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_properties.tabColor = NAVY

    for row in ws1.iter_rows(min_row=1, max_row=20, min_col=1, max_col=6):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=CREAM)

    ws1["B2"] = "Stufly Price Tracker Report"
    ws1["B2"].font = Font(name="Arial", size=16, bold=True, color=NAVY)
    ws1["B3"] = f"{origin} -> {destination}  |  Tracked {sorted_dates[0].strftime('%d %b %Y') if sorted_dates else '-'} - {sorted_dates[-1].strftime('%d %b %Y') if sorted_dates else '-'}"
    ws1["B3"].font = Font(name="Arial", size=11, color="7A7362")

    thin = Side(style="thin", color="D8D2C2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    label_fill = PatternFill("solid", fgColor=CREAM_DIM)

    rows = [
        ("Airlines tracked", ", ".join(airlines)),
        ("Cheapest price found", f"Rs {best_price:,}" if best_price else "No prices recorded"),
        ("Cheapest airline", best_airline or "-"),
        ("Cheapest day", best_date.strftime("%d %b %Y") if best_date else "-"),
    ]
    for i, (label, value) in enumerate(rows):
        r = 6 + i
        ws1[f"B{r}"] = label
        ws1[f"B{r}"].fill = label_fill
        ws1[f"B{r}"].font = Font(name="Arial", size=10, bold=True, color="5C5646")
        ws1[f"B{r}"].border = border
        ws1[f"D{r}"] = value
        ws1[f"D{r}"].border = border
        ws1[f"D{r}"].font = Font(name="Arial", size=12, bold=True)
        if i == 1:
            ws1[f"D{r}"].fill = PatternFill("solid", fgColor=SAGE)

    ws1.column_dimensions["B"].width = 26
    ws1.column_dimensions["D"].width = 30

    ws2 = wb.create_sheet("Daily Prices")
    ws2.sheet_properties.tabColor = GOLD
    for row in ws2.iter_rows(min_row=1, max_row=30, min_col=1, max_col=8):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=CREAM)

    ws2["B2"] = "Daily Price Tracking"
    ws2["B2"].font = Font(name="Arial", size=16, bold=True, color=NAVY)

    headers = ["Date"] + airlines
    for i, h in enumerate(headers):
        c = ws2.cell(row=4, column=2 + i, value=h)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center")
        c.border = border

    for i, d in enumerate(sorted_dates):
        r = 5 + i
        ws2.cell(row=r, column=2, value=d.strftime("%d %b %Y")).font = Font(name="Arial", size=10, bold=True, color=NAVY)
        for j, airline in enumerate(airlines):
            price = by_date[d].get(airline)
            cell = ws2.cell(row=r, column=3 + j, value=price if price is not None else "-")
            if price is not None:
                cell.number_format = '"Rs"#,##0'
            cell.border = border
            if price is not None and price == best_price:
                cell.fill = PatternFill("solid", fgColor=SAGE)
                cell.font = Font(name="Arial", size=10, bold=True, color="2F6B54")

    discount_row = 6 + len(sorted_dates)
    ws2[f"B{discount_row}"] = "Student discounts"
    ws2[f"B{discount_row}"].font = Font(name="Arial", size=11, bold=True, color=NAVY)
    for i, airline in enumerate(airlines):
        programs = get_discounts_for_airline(airline)
        pct_programs = [p for p in programs if p["discount_type"] == "percentage"]
        if pct_programs:
            best_pct = max(p["discount_value"] for p in pct_programs)
            note = f"{airline}: up to {best_pct}% off"
        elif programs:
            note = f"{airline}: baggage/other perk, no fare discount"
        else:
            note = f"{airline}: no known student program"
        ws2[f"B{discount_row + 1 + i}"] = note
        ws2[f"B{discount_row + 1 + i}"].font = Font(name="Arial", size=9, color="5C5646")

    link_row = discount_row + len(airlines) + 2
    ws2[f"B{link_row}"] = "Verify current price & book"
    ws2[f"B{link_row}"].font = Font(name="Arial", size=11, bold=True, color=NAVY)
    ws2[f"B{link_row + 1}"] = f"Search {origin} to {destination} on Stufly"
    ws2[f"B{link_row + 1}"].hyperlink = f"https://stufly.app/search?origin={origin}&destination={destination}"
    ws2[f"B{link_row + 1}"].font = Font(name="Arial", size=10, underline="single", color="1155CC")

    for col, w in zip("BCDEF", [16, 15, 15, 15, 15]):
        ws2.column_dimensions[col].width = w

    wb.save(filepath)


def _build_html_email(tracker: dict, history: list[dict], best_price, best_airline, best_date) -> str:
    """
    Builds the styled HTML email body -- this replicates the preview we
    designed and approved earlier, just filled in with real numbers from
    this specific tracker instead of mock data.
    """
    origin, destination = tracker["origin"], tracker["destination"]
    airlines = tracker["airlines"].split(",")

    airline_lows = {}
    for row in history:
        if row["price"] is None:
            continue
        a = row["airline"]
        if a not in airline_lows or row["price"] < airline_lows[a]["price"]:
            airline_lows[a] = {"price": row["price"], "date": row["checked_date"]}

    table_rows = ""
    for airline in airlines:
        info = airline_lows.get(airline)
        is_best = info and best_price is not None and info["price"] == best_price
        row_style = "background:#D9F0E6;font-weight:700;color:#2F6B54;" if is_best else ""
        price_text = f"Rs {info['price']:,}" if info else "No data"
        date_text = info["date"].strftime("%d %b") if info else "-"
        table_rows += f"""
          <tr style="{row_style}">
            <td style="padding:9px 12px;border-bottom:1px solid #F0EBDD;">{airline}</td>
            <td style="padding:9px 12px;border-bottom:1px solid #F0EBDD;text-align:right;font-family:'Courier New',monospace;">{price_text}</td>
            <td style="padding:9px 12px;border-bottom:1px solid #F0EBDD;">{date_text}</td>
          </tr>"""

    headline = f"<b style='color:#C98F1F;'>{best_airline}</b> came in cheapest at <b style='color:#C98F1F;'>Rs {best_price:,}</b>, on {best_date.strftime('%d %b')}." if best_price else "We couldn't find prices for these airlines during this window."

    return f"""
    <!DOCTYPE html><html><body style="margin:0;padding:0;background:#E8E8E8;font-family:Arial,sans-serif;">
    <div style="max-width:600px;margin:20px auto;background:#fff;border-radius:10px;overflow:hidden;">
      <div style="background:linear-gradient(135deg,#14213D,#0D1730);padding:24px 28px;color:#FDF8F0;">
        <div style="font-family:Georgia,serif;font-size:20px;font-weight:700;color:#FFB627;">Stufly</div>
        <div style="font-size:11px;letter-spacing:0.1em;text-transform:uppercase;color:rgba(253,248,240,0.5);margin-top:2px;">
          {tracker['duration_days']}-day price tracker &middot; complete
        </div>
      </div>
      <div style="padding:26px 28px 8px;">
        <p style="font-size:15px;color:#1B1F2E;margin-bottom:14px;">Hi,</p>
        {"<span style='display:inline-block;background:#D9F0E6;color:#2F6B54;font-weight:700;font-size:11.5px;letter-spacing:0.06em;text-transform:uppercase;padding:6px 14px;border-radius:100px;margin-bottom:16px;'>Best deal found</span><br>" if best_price else ""}
        <h1 style="font-family:Georgia,serif;font-size:21px;color:#14213D;line-height:1.4;margin:10px 0 6px;">{headline}</h1>
        <p style="font-size:13.5px;color:#7A7362;margin-bottom:22px;">
          You asked us to watch {origin} to {destination} across {', '.join(airlines)} for {tracker['duration_days']} days.
          Here's how they compared &mdash; full daily breakdown in the attached sheet.
        </p>
        <table style="width:100%;border-collapse:collapse;margin-bottom:22px;font-size:13px;">
          <tr>
            <th style="background:#14213D;color:#FDF8F0;text-align:left;padding:9px 12px;font-size:11px;letter-spacing:0.05em;text-transform:uppercase;">Airline</th>
            <th style="background:#14213D;color:#FDF8F0;text-align:right;padding:9px 12px;font-size:11px;letter-spacing:0.05em;text-transform:uppercase;">Lowest seen</th>
            <th style="background:#14213D;color:#FDF8F0;text-align:left;padding:9px 12px;font-size:11px;letter-spacing:0.05em;text-transform:uppercase;">On</th>
          </tr>
          {table_rows}
        </table>
      </div>
      <div style="text-align:center;padding:6px 0 28px;">
        <a href="https://stufly.app/search?origin={origin}&destination={destination}"
           style="display:inline-block;background:#14213D;color:#FDF8F0;text-decoration:none;font-weight:600;font-size:14px;padding:13px 34px;border-radius:8px;">
          Search this route on Stufly
        </a>
      </div>
      <div style="background:#FDF8F0;border:1px dashed #E0DACB;border-radius:10px;margin:0 28px 26px;padding:14px 16px;font-size:12.5px;color:#7A7362;">
        The attached Excel has a day-by-day table for all airlines you tracked, plus a summary tab with the overall cheapest option highlighted, student discount notes, and links to verify and book.
      </div>
      <div style="padding:18px 28px 28px;font-size:11px;color:#B4AC98;border-top:1px solid #F0EBDD;">
        You're getting this because you set up a price tracker on Stufly for {origin} to {destination}.
      </div>
    </div>
    </body></html>
    """


def send_tracker_report(tracker: dict):
    """
    Puts together the finished report and emails it -- this is where a
    tracker's whole story comes together: pull the recorded history, build
    the spreadsheet, build the styled HTML email around it, and send both.

    If Gmail isn't configured, we log and quietly skip rather than
    crashing the poller -- a missing email credential shouldn't take down
    the whole background price-tracking process.
    """
    if not _GMAIL_ADDRESS or not _GMAIL_APP_PASSWORD:
        print(f"[notifications] Gmail not configured -- skipping report for tracker {tracker['id']}")
        return False

    history = get_tracker_history(tracker["id"])
    filepath = f"/tmp/stufly_tracker_{tracker['id']}.xlsx"
    _build_report_excel(tracker, history, filepath)

    best_price, best_airline, best_date = None, None, None
    for row in history:
        if row["price"] is not None and (best_price is None or row["price"] < best_price):
            best_price, best_airline, best_date = row["price"], row["airline"], row["checked_date"]

    subject = f"Your {tracker['origin']} to {tracker['destination']} tracker: results are in"
    html_body = _build_html_email(tracker, history, best_price, best_airline, best_date)

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"] = f"Stufly Price Alerts <{_GMAIL_ADDRESS}>"
    msg["To"] = tracker["email"]

    # Attach the HTML body as its own "alternative" part -- some email
    # clients don't render HTML, so this is the standard way to give
    # them a real body instead of a blank email. We don't bother with a
    # separate plain-text fallback here since the HTML is simple enough
    # that every modern client renders it fine.
    msg.attach(MIMEText(html_body, "html"))

    with open(filepath, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename=stufly_tracker_{tracker['origin']}_{tracker['destination']}.xlsx")
    msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(_GMAIL_ADDRESS, _GMAIL_APP_PASSWORD)
            server.sendmail(_GMAIL_ADDRESS, [tracker["email"]], msg.as_string())
        print(f"[notifications] Sent tracker report to {tracker['email']} for tracker {tracker['id']}")
        return True
    except Exception as e:
        print(f"[notifications] Failed to send tracker report {tracker['id']}: {e}")
        return False